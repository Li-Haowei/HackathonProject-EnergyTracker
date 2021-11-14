# -*- coding: utf-8 -*-
"""
Created on Sat Nov 13 12:26:34 2021

@author: Haowei Li
"""
import openpyxl
path = "./dataBase.xlsx"

class dataBase():
     def __init__(self):
          self.base = openpyxl.load_workbook(path)
          self.emotionTracker = self.base['emotionTracker']
          self.assignmentTracker = self.base['assignmentTracker']
     
     def addEmotion(self,date,level,energy,focus):
          """add the date and emotion level into dataBase"""
          i = 0
          cell_val = ''
          # Finds which row is blank first
          while cell_val != None:
               i += 1
               cell_val = self.emotionTracker['A' + str(i)].value         
          #x = input('Prompt: ')
          self.emotionTracker['A' + str(i)] = date
          self.emotionTracker['B' + str(i)] = level
          self.emotionTracker['C' + str(i)] = energy
          self.emotionTracker['D' + str(i)] = focus
          self.emotionTracker['L1'].value += 1
          self.base.save('dataBase.xlsx')
          
     def addAssignment(self,deadline,course,percentage,duration):
          """add the deadline, course, percentage of assignment, and finish status into dataBase"""
          i = 0
          cell_val = ''
          # Finds which row is blank first
          while cell_val != None:
               i += 1
               cell_val = self.assignmentTracker['A' + str(i)].value         
          self.assignmentTracker['A' + str(i)] = deadline
          self.assignmentTracker['B' + str(i)] = course
          self.assignmentTracker['C' + str(i)] = percentage
          self.assignmentTracker['D' + str(i)] = duration
          self.assignmentTracker['L1'].value += 1
          self.base.save('dataBase.xlsx')
           
     def delEmotion(self,date):
          """delete the emotion records on that date"""
          i = 1
          cell_val = self.emotionTracker.cell(row=i,column=1)
          while cell_val.value!=date and i<=self.emotionTracker['L1'].value:
               i += 1
               cell_val = self.emotionTracker.cell(row=i,column=1)
          if i>self.emotionTracker['L1'].value:
               return False
          self.emotionTracker['A' + str(i)] = None
          self.emotionTracker['B' + str(i)] = None
          self.emotionTracker['C' + str(i)] = None
          self.emotionTracker['D' + str(i)] = None
          self.base.save('dataBase.xlsx')
          
          return True
     def delAssignment(self,date):
          """delete the assignment records on that date"""
          i = 1
          cell_val = self.assignmentTracker.cell(row=i,column=1)
          while cell_val.value!=date and i<=self.assignmentTracker['L1'].value:
               i += 1
               cell_val = self.assignmentTracker.cell(row=i,column=1)
          if i>self.assignmentTracker['L1'].value:
               return False
          self.assignmentTracker['A' + str(i)] = None
          self.assignmentTracker['B' + str(i)] = None
          self.assignmentTracker['C' + str(i)] = None
          self.assignmentTracker['D' + str(i)] = None
          self.base.save('dataBase.xlsx')
          
          return True
     def getEmotion(self):
          result = []
          i = 1
          cell_val = self.emotionTracker.cell(row=i,column=1)
          while i<=self.emotionTracker['L1'].value:
               if cell_val.value != None:
                    result.append(
                              (self.emotionTracker.cell(row=i,column=1).value,
                               self.emotionTracker.cell(row=i,column=2).value,
                               self.emotionTracker.cell(row=i,column=3).value,
                               self.emotionTracker.cell(row=i,column=4).value)
                              )
               i += 1
               cell_val = self.emotionTracker.cell(row=i,column=1)
          return result
     
     def getAssignment(self):
          result = []
          i = 1
          cell_val = self.assignmentTracker.cell(row=i,column=1)
          while i<=self.assignmentTracker['L1'].value:
               if cell_val.value != None:
                    result.append((
                              self.assignmentTracker.cell(row=i,column=1).value,
                               self.assignmentTracker.cell(row=i,column=2).value,
                               self.assignmentTracker.cell(row=i,column=3).value,
                               self.assignmentTracker.cell(row=i,column=4).value
                               ))
               i += 1
               cell_val = self.assignmentTracker.cell(row=i,column=1)
          return result
     
     def printEmotion(self):
          i = 1
          cell_val = self.emotionTracker.cell(row=i,column=1)
          while i<=self.emotionTracker['L1'].value:
               if cell_val.value != None:
                    print(
                    "Date: ", self.emotionTracker.cell(row=i,column=1).value,
                    " Emotion Level: ",self.emotionTracker.cell(row=i,column=2).value,
                    " Energy Level: ",self.emotionTracker.cell(row=i,column=2).value,
                    " Focus Level: ",self.emotionTracker.cell(row=i,column=2).value
                    )
               
               i += 1
               cell_val = self.emotionTracker.cell(row=i,column=1)
     
     def printAssignment(self):
          i = 1
          cell_val = self.assignmentTracker.cell(row=i,column=1)
          while i<=self.assignmentTracker['L1'].value:
               if cell_val.value != None:
                    print(
                         "Deadline: ", self.assignmentTracker.cell(row=i,column=1).value,
                         "Course: ", self.assignmentTracker.cell(row=i,column=2).value,
                         "Percentage: ",self.assignmentTracker.cell(row=i,column=3).value,
                         "Duration: ",self.assignmentTracker.cell(row=i,column=4).value
                         )
               i += 1
               cell_val = self.assignmentTracker.cell(row=i,column=1)
               
     def delEmotionDataBase(self):
          """delete all the emotion records on that date"""
          i = 1
          while i<=self.emotionTracker['L1'].value:
               self.emotionTracker['A' + str(i)] = None
               self.emotionTracker['B' + str(i)] = None
               self.emotionTracker['C' + str(i)] = None
               self.emotionTracker['D' + str(i)] = None
               i += 1
          self.emotionTracker['L1'].value = 0
          self.base.save('dataBase.xlsx')
     def delAssignmentDataBase(self):
          """delete all the emotion records on that date"""
          i = 1
          while i<=self.assignmentTracker['L1'].value:
               self.assignmentTracker['A' + str(i)] = None
               self.assignmentTracker['B' + str(i)] = None
               self.assignmentTracker['C' + str(i)] = None
               self.assignmentTracker['D' + str(i)] = None
               i += 1
          self.assignmentTracker['L1'].value = 0
          self.base.save('dataBase.xlsx')
     
     
